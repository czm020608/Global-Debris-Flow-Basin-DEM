#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
Compute main-channel and station metrics from DEM hydrology rasters.

Expected point TXT/CSV formats:
    start/outlet: ID, X, Y
    stations:     ID, X, Y, Z(optional)

Coordinates must be in the DEM CRS, e.g. EPSG:32648 for this project.
Use prepare_google_points.py to convert Google Maps lon/lat first.
"""

from __future__ import annotations

import argparse
import math
from dataclasses import dataclass
from pathlib import Path

import numpy as np
import pandas as pd
import rasterio
from openpyxl import load_workbook
from openpyxl.chart import LineChart, Reference
from openpyxl.styles import Alignment, Font, PatternFill
from rasterio.transform import rowcol, xy


WHITEBOX_D8 = {
    1: (0, 1),
    2: (-1, 1),
    4: (-1, 0),
    8: (-1, -1),
    16: (0, -1),
    32: (1, -1),
    64: (1, 0),
    128: (1, 1),
}

ESRI_D8 = {
    1: (0, 1),
    2: (1, 1),
    4: (1, 0),
    8: (1, -1),
    16: (0, -1),
    32: (-1, -1),
    64: (-1, 0),
    128: (-1, 1),
}


@dataclass(frozen=True)
class PointRecord:
    point_id: str
    x: float
    y: float
    z: float | None = None


def read_points(path: Path) -> list[PointRecord]:
    sep = "\t" if path.suffix.lower() == ".txt" else ","
    df = pd.read_csv(path, sep=sep)
    df.columns = [str(c).strip() for c in df.columns]
    required = {"ID", "X", "Y"}
    missing = required.difference(df.columns)
    if missing:
        raise ValueError(f"{path} is missing required column(s): {', '.join(sorted(missing))}")

    z_values = df["Z"] if "Z" in df.columns else [None] * len(df)
    return [
        PointRecord(str(row.ID), float(row.X), float(row.Y), None if pd.isna(z) else float(z))
        for row, z in zip(df.itertuples(index=False), z_values)
    ]


def validate_projected_points(points: list[PointRecord], transform, shape: tuple[int, int], label: str) -> None:
    rows, cols = shape
    outside = []
    for point in points:
        row, col = rowcol(transform, point.x, point.y)
        if row < 0 or row >= rows or col < 0 or col >= cols:
            outside.append(point.point_id)
    if outside:
        raise ValueError(
            f"{label} point(s) outside DEM bounds after projection: {', '.join(outside)}. "
            "Check that the point CRS matches the DEM analysis CRS."
        )


def nearest_high_acc_cell(flow_acc: np.ndarray, transform, point: PointRecord, radius_m: float) -> tuple[int, int]:
    row, col = rowcol(transform, point.x, point.y)
    radius_cells = max(1, int(math.ceil(radius_m / abs(transform.a))))
    r0 = max(0, row - radius_cells)
    r1 = min(flow_acc.shape[0] - 1, row + radius_cells)
    c0 = max(0, col - radius_cells)
    c1 = min(flow_acc.shape[1] - 1, col + radius_cells)
    window = flow_acc[r0 : r1 + 1, c0 : c1 + 1]
    if window.size == 0:
        raise ValueError(f"Point {point.point_id} is outside raster bounds.")
    local = np.unravel_index(np.nanargmax(window), window.shape)
    return r0 + int(local[0]), c0 + int(local[1])


def trace_downstream(
    start_rc: tuple[int, int],
    outlet_rc: tuple[int, int],
    flow_dir: np.ndarray,
    pointer_scheme: str,
    stop_distance_cells: int = 2,
    max_steps: int = 200000,
) -> list[tuple[int, int]]:
    mapping = WHITEBOX_D8 if pointer_scheme == "whitebox" else ESRI_D8
    rows, cols = flow_dir.shape
    path = [start_rc]
    seen = {start_rc}
    r, c = start_rc

    for _ in range(max_steps):
        if abs(r - outlet_rc[0]) <= stop_distance_cells and abs(c - outlet_rc[1]) <= stop_distance_cells:
            break
        direction = int(flow_dir[r, c])
        if direction == 0:
            break
        if direction not in mapping:
            raise ValueError(f"Unexpected D8 direction value {direction} at row={r}, col={c}")
        dr, dc = mapping[direction]
        nr, nc = r + dr, c + dc
        if nr < 0 or nr >= rows or nc < 0 or nc >= cols:
            break
        if (nr, nc) in seen:
            raise ValueError(f"Flow path loop detected at row={nr}, col={nc}")
        path.append((nr, nc))
        seen.add((nr, nc))
        r, c = nr, nc
    return path


def trace_downstream_auto(
    start_rc: tuple[int, int],
    outlet_rc: tuple[int, int],
    flow_dir: np.ndarray,
    pointer_scheme: str,
) -> tuple[list[tuple[int, int]], str]:
    schemes = ["whitebox", "esri"] if pointer_scheme == "auto" else [pointer_scheme]
    candidates = []
    errors = []
    for scheme in schemes:
        try:
            path = trace_downstream(start_rc, outlet_rc, flow_dir, scheme)
            last = path[-1]
            outlet_gap = math.hypot(last[0] - outlet_rc[0], last[1] - outlet_rc[1])
            candidates.append((outlet_gap, -len(path), path, scheme))
        except Exception as exc:
            errors.append(f"{scheme}: {exc}")
    if not candidates:
        raise ValueError("No valid downstream trace found. " + " | ".join(errors))
    candidates.sort(key=lambda item: (item[0], item[1]))
    _, _, path, scheme = candidates[0]
    return path, scheme


def path_dataframe(path: list[tuple[int, int]], dem: np.ndarray, flow_acc: np.ndarray, transform) -> pd.DataFrame:
    records = []
    cumulative = 0.0
    previous_xy = None
    previous_z = None
    for idx, (r, c) in enumerate(path):
        x, y = xy(transform, r, c, offset="center")
        z = float(dem[r, c])
        if previous_xy is None:
            step = 0.0
            slope = np.nan
        else:
            step = math.hypot(x - previous_xy[0], y - previous_xy[1])
            cumulative += step
            slope = (previous_z - z) / step if step else np.nan
        records.append(
            {
                "seq": idx,
                "row": r,
                "col": c,
                "distance_m": cumulative,
                "X": x,
                "Y": y,
                "elevation_m": z,
                "step_length_m": step,
                "segment_slope": slope,
                "flow_acc_cells": float(flow_acc[r, c]),
            }
        )
        previous_xy = (x, y)
        previous_z = z
    return pd.DataFrame.from_records(records)


def interpolate_elevation(profile: pd.DataFrame, distance_m: float) -> float:
    distances = profile["distance_m"].to_numpy()
    elevations = profile["elevation_m"].to_numpy()
    distance_m = float(np.clip(distance_m, distances.min(), distances.max()))
    return float(np.interp(distance_m, distances, elevations))


def slope_around(profile: pd.DataFrame, distance_m: float, half_window_m: float) -> float:
    d0 = max(0.0, distance_m - half_window_m)
    d1 = min(float(profile["distance_m"].iloc[-1]), distance_m + half_window_m)
    if d1 <= d0:
        return np.nan
    return (interpolate_elevation(profile, d0) - interpolate_elevation(profile, d1)) / (d1 - d0)


def nearest_profile_index(profile: pd.DataFrame, x: float, y: float) -> int:
    dx = profile["X"].to_numpy() - x
    dy = profile["Y"].to_numpy() - y
    return int(np.argmin(dx * dx + dy * dy))


def estimate_cross_section_width(
    profile: pd.DataFrame,
    idx: int,
    dem: np.ndarray,
    transform,
    bank_relief_m: float,
    max_width_m: float,
) -> float:
    if len(profile) < 3:
        return np.nan

    i0 = max(0, idx - 1)
    i1 = min(len(profile) - 1, idx + 1)
    dx = float(profile.loc[i1, "X"] - profile.loc[i0, "X"])
    dy = float(profile.loc[i1, "Y"] - profile.loc[i0, "Y"])
    length = math.hypot(dx, dy)
    if length == 0:
        return np.nan

    nx, ny = -dy / length, dx / length
    x0 = float(profile.loc[idx, "X"])
    y0 = float(profile.loc[idx, "Y"])
    z0 = float(profile.loc[idx, "elevation_m"])
    cell = abs(transform.a)
    max_steps = int(max_width_m / (2 * cell))

    widths = []
    for sign in (-1, 1):
        side_width = max_width_m / 2
        for step in range(1, max_steps + 1):
            x = x0 + sign * nx * step * cell
            y = y0 + sign * ny * step * cell
            try:
                r, c = rowcol(transform, x, y)
            except Exception:
                side_width = (step - 1) * cell
                break
            if r < 0 or r >= dem.shape[0] or c < 0 or c >= dem.shape[1]:
                side_width = (step - 1) * cell
                break
            if float(dem[r, c]) - z0 >= bank_relief_m:
                side_width = step * cell
                break
        widths.append(side_width)
    return float(sum(widths))


def station_metrics(
    stations: list[PointRecord],
    profile: pd.DataFrame,
    dem: np.ndarray,
    transform,
    slope_window_m: float,
    bank_relief_m: float,
    max_width_m: float,
    station_offset_review_m: float,
) -> pd.DataFrame:
    rows = []
    for station in stations:
        idx = nearest_profile_index(profile, station.x, station.y)
        nearest = profile.loc[idx]
        offset = math.hypot(float(nearest["X"]) - station.x, float(nearest["Y"]) - station.y)
        rows.append(
            {
                "station_id": station.point_id,
                "station_X": station.x,
                "station_Y": station.y,
                "station_Z_input": station.z,
                "nearest_channel_distance_m": float(nearest["distance_m"]),
                "nearest_channel_X": float(nearest["X"]),
                "nearest_channel_Y": float(nearest["Y"]),
                "nearest_channel_elevation_m": float(nearest["elevation_m"]),
                "station_to_channel_offset_m": offset,
                "needs_manual_review": "yes" if offset > station_offset_review_m else "no",
                "slope_100m_before_after": slope_around(profile, float(nearest["distance_m"]), slope_window_m),
                "dem_cross_section_width_m": estimate_cross_section_width(
                    profile, idx, dem, transform, bank_relief_m, max_width_m
                ),
            }
        )
    return pd.DataFrame.from_records(rows)


def summary_metrics(profile: pd.DataFrame, cell_area_m2: float) -> pd.DataFrame:
    slopes = profile["segment_slope"].replace([np.inf, -np.inf], np.nan)
    steep_idx = int(slopes.idxmax()) if slopes.notna().any() else 0
    gentle_idx = int(slopes.idxmin()) if slopes.notna().any() else 0
    area_m2 = float(profile["flow_acc_cells"].max() * cell_area_m2)
    records = [
        ("channel_length_m", float(profile["distance_m"].iloc[-1])),
        ("channel_drop_m", float(profile["elevation_m"].iloc[0] - profile["elevation_m"].iloc[-1])),
        ("average_channel_slope", slope_around(profile, 0.5 * float(profile["distance_m"].iloc[-1]), 1e12)),
        ("max_segment_slope", float(slopes.loc[steep_idx])),
        ("max_slope_distance_m", float(profile.loc[steep_idx, "distance_m"])),
        ("min_segment_slope", float(slopes.loc[gentle_idx])),
        ("min_slope_distance_m", float(profile.loc[gentle_idx, "distance_m"])),
        ("drainage_area_m2_from_flow_acc", area_m2),
        ("drainage_area_km2_from_flow_acc", area_m2 / 1e6),
    ]
    return pd.DataFrame(records, columns=["metric", "value"])


def validation_metrics(
    dem_crs: str,
    cellsize_x: float,
    cellsize_y: float,
    start_rc: tuple[int, int],
    outlet_rc: tuple[int, int],
    traced_outlet_rc: tuple[int, int],
    selected_scheme: str,
    snap_radius_m: float,
    slope_window_m: float,
    bank_relief_m: float,
    max_width_m: float,
    station_offset_review_m: float,
) -> pd.DataFrame:
    outlet_gap_cells = math.hypot(traced_outlet_rc[0] - outlet_rc[0], traced_outlet_rc[1] - outlet_rc[1])
    mean_cellsize = (abs(cellsize_x) + abs(cellsize_y)) / 2
    records = [
        ("analysis_crs", dem_crs),
        ("cellsize_x_m", abs(cellsize_x)),
        ("cellsize_y_m", abs(cellsize_y)),
        ("selected_d8_pointer_scheme", selected_scheme),
        ("start_row_col", f"{start_rc[0]},{start_rc[1]}"),
        ("outlet_row_col", f"{outlet_rc[0]},{outlet_rc[1]}"),
        ("traced_end_row_col", f"{traced_outlet_rc[0]},{traced_outlet_rc[1]}"),
        ("traced_end_to_outlet_gap_m", outlet_gap_cells * mean_cellsize),
        ("snap_radius_m", snap_radius_m),
        ("station_slope_half_window_m", slope_window_m),
        ("bank_relief_threshold_m", bank_relief_m),
        ("max_width_search_m", max_width_m),
        ("station_offset_review_threshold_m", station_offset_review_m),
    ]
    return pd.DataFrame(records, columns=["metric", "value"])


def write_excel(output_path: Path, summary: pd.DataFrame, profile: pd.DataFrame, stations: pd.DataFrame) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        summary.to_excel(writer, sheet_name="Summary", index=False)
        profile.to_excel(writer, sheet_name="Channel_Profile", index=False)
        stations.to_excel(writer, sheet_name="Station_Metrics", index=False)

    wb = load_workbook(output_path)
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    for ws in wb.worksheets:
        ws.freeze_panes = "A2"
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
        for col in ws.columns:
            max_len = max(len(str(cell.value)) if cell.value is not None else 0 for cell in col)
            ws.column_dimensions[col[0].column_letter].width = min(max(max_len + 2, 12), 32)

    if "Channel_Profile" in wb.sheetnames and wb["Channel_Profile"].max_row > 2:
        ws = wb["Channel_Profile"]
        chart = LineChart()
        chart.title = "Main Channel Longitudinal Profile"
        chart.y_axis.title = "Elevation (m)"
        chart.x_axis.title = "Profile point sequence"
        data = Reference(ws, min_col=7, min_row=1, max_row=ws.max_row)
        chart.add_data(data, titles_from_data=True)
        ws.add_chart(chart, "M2")

    wb.save(output_path)


def compute_metrics(
    dem_path: Path,
    flow_dir_path: Path,
    flow_acc_path: Path,
    start_path: Path,
    outlet_path: Path,
    stations_path: Path | None,
    output_path: Path,
    snap_radius_m: float,
    pointer_scheme: str,
    slope_window_m: float,
    bank_relief_m: float,
    max_width_m: float,
    station_offset_review_m: float,
) -> Path:
    starts = read_points(start_path)
    outlets = read_points(outlet_path)
    station_points = read_points(stations_path) if stations_path else []
    if len(starts) != 1 or len(outlets) != 1:
        raise ValueError("Start and outlet files must each contain exactly one point.")

    with rasterio.open(dem_path) as dem_src, rasterio.open(flow_dir_path) as dir_src, rasterio.open(flow_acc_path) as acc_src:
        if dem_src.crs != dir_src.crs or dem_src.crs != acc_src.crs:
            raise ValueError("DEM, flow_dir and flow_acc CRS do not match.")
        if dem_src.transform != dir_src.transform or dem_src.transform != acc_src.transform:
            raise ValueError("DEM, flow_dir and flow_acc transforms do not match.")
        if dem_src.crs and dem_src.crs.to_epsg() == 3857:
            raise ValueError("EPSG:3857 Web Mercator is not allowed for length/area/slope analysis.")
        validate_projected_points(starts, dem_src.transform, (dem_src.height, dem_src.width), "Start")
        validate_projected_points(outlets, dem_src.transform, (dem_src.height, dem_src.width), "Outlet")
        validate_projected_points(station_points, dem_src.transform, (dem_src.height, dem_src.width), "Station")
        dem = dem_src.read(1)
        flow_dir = dir_src.read(1)
        flow_acc = acc_src.read(1)
        transform = dem_src.transform
        dem_crs = str(dem_src.crs)
        cellsize_x, cellsize_y = dem_src.res
        cell_area_m2 = abs(transform.a * transform.e)

    start_rc = nearest_high_acc_cell(flow_acc, transform, starts[0], snap_radius_m)
    outlet_rc = nearest_high_acc_cell(flow_acc, transform, outlets[0], snap_radius_m)
    path, selected_scheme = trace_downstream_auto(start_rc, outlet_rc, flow_dir, pointer_scheme)
    profile = path_dataframe(path, dem, flow_acc, transform)

    if len(profile) < 2:
        raise ValueError("The traced channel contains fewer than two cells. Check point locations and D8 pointer scheme.")

    summary = summary_metrics(profile, cell_area_m2)
    validation = validation_metrics(
        dem_crs,
        cellsize_x,
        cellsize_y,
        start_rc,
        outlet_rc,
        path[-1],
        selected_scheme,
        snap_radius_m,
        slope_window_m,
        bank_relief_m,
        max_width_m,
        station_offset_review_m,
    )
    summary = pd.concat(
        [
            validation,
            pd.DataFrame([("traced_cells", int(len(profile)))], columns=["metric", "value"]),
            summary,
        ],
        ignore_index=True,
    )
    if stations_path:
        stations = station_metrics(
            station_points,
            profile,
            dem,
            transform,
            slope_window_m,
            bank_relief_m,
            max_width_m,
            station_offset_review_m,
        )
    else:
        stations = pd.DataFrame(
            columns=[
                "station_id",
                "station_X",
                "station_Y",
                "station_Z_input",
                "nearest_channel_distance_m",
                "nearest_channel_X",
                "nearest_channel_Y",
                "nearest_channel_elevation_m",
                "station_to_channel_offset_m",
                "needs_manual_review",
                "slope_100m_before_after",
                "dem_cross_section_width_m",
            ]
        )
    write_excel(output_path, summary, profile, stations)
    return output_path


def main() -> None:
    parser = argparse.ArgumentParser(description="Compute DEM-derived channel and station metrics.")
    parser.add_argument("--dem", default=Path("raster/jiangjiagou.tif"), type=Path)
    parser.add_argument("--flow-dir", default=Path("temp_working_dir/flow_dir.tif"), type=Path)
    parser.add_argument("--flow-acc", default=Path("temp_working_dir/flow_acc.tif"), type=Path)
    parser.add_argument("--start", required=True, type=Path, help="TXT/CSV with one main-channel start point: ID,X,Y")
    parser.add_argument("--outlet", required=True, type=Path, help="TXT/CSV with one outlet point: ID,X,Y")
    parser.add_argument("--stations", type=Path, help="TXT/CSV station file: ID,X,Y,Z(optional)")
    parser.add_argument("--output", default=Path("outputs/channel_metrics.xlsx"), type=Path)
    parser.add_argument("--snap-radius-m", default=100.0, type=float)
    parser.add_argument("--pointer-scheme", choices=["auto", "whitebox", "esri"], default="auto")
    parser.add_argument("--slope-window-m", default=100.0, type=float)
    parser.add_argument("--bank-relief-m", default=5.0, type=float)
    parser.add_argument("--max-width-m", default=300.0, type=float)
    parser.add_argument("--station-offset-review-m", default=100.0, type=float)
    args = parser.parse_args()

    output = compute_metrics(
        args.dem,
        args.flow_dir,
        args.flow_acc,
        args.start,
        args.outlet,
        args.stations,
        args.output,
        args.snap_radius_m,
        args.pointer_scheme,
        args.slope_window_m,
        args.bank_relief_m,
        args.max_width_m,
        args.station_offset_review_m,
    )
    print(f"Wrote {output}")


if __name__ == "__main__":
    main()
